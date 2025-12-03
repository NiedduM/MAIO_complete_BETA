
# -*- coding: utf-8 -*-
"""
Created on Sun Feb 25 19:56:41 2024

@author: marce
"""

import Model as md
import numpy as np
import matplotlib.pyplot as plt
import os as os
import openpyxl as opxl
import pickle as pk
import math as math

class Experimenter(object):
    
    
    def __init__(self):
        
        self.model = object()
        
        return
    
    def create_initcond_simplified(self,seed, Y, GDP_RoW, RoW_price, c_RoW, alpha_RoW, n_sector, C, delta, gamma, cap_ut ,cap_target_speed, KAU_price, wages, labor_tech_coeff , KAU_from_RoW, n_households, lambdaT, csi, alpha, H_from_RoW, alpha_gov ,Gov_from_RoW, tax_rates, ub_frac, transfer_frac, n_indep_periods =100, shocks = {'active': False}):


        
        A_mod = np.zeros((n_sector, n_sector))
        
        for i in range(n_sector):
            for j in range(n_sector):
                
                A_mod[i,j] = C[i,j]
                
                if(cap_ut[i,j] >0):
                    A_mod[i,j] += delta[i,j]*gamma[i,j]    # /cap_ut[i,j] uncomment if you want depreciation proportional to the stock level
                
                A_mod[i,j] = KAU_price[i]*A_mod[i,j]/KAU_price[j]
                
        a_KAU = np.zeros(n_sector)
        a_H = np.sum(H_from_RoW*alpha)
        a_G = np.sum(Gov_from_RoW*alpha_gov)
        
        for i in range(n_sector):            
            for j in range(n_sector):
                
                a_KAU[i] += RoW_price[j]/KAU_price[j,0]*KAU_from_RoW[j,i]*A_mod[j,i]
        
        A_eff = np.zeros((n_sector, n_sector))
        A_cost =  np.zeros((n_sector, n_sector))
        
        alpha_new = (1-H_from_RoW)*alpha + alpha_RoW*a_H
        alpha_gov_new = (1-Gov_from_RoW)*alpha_gov + alpha_RoW*a_G    
        
        
        for i in range(n_sector):            
            for j in range(n_sector): 
                
                A_eff[i,j] = (1-KAU_from_RoW[i,j])*A_mod[i,j] + alpha_RoW[i]*a_KAU[j]
                A_cost[i,j] = (1+ (RoW_price[i]/KAU_price[i] -1)*KAU_from_RoW[i,j])*A_mod[i,j]

        L_eff = np.linalg.inv(np.eye(n_sector) - A_eff)
        
        A_d = a_G + np.dot(a_KAU.T, np.dot(L_eff, alpha_gov_new) )
        B_d = a_H - a_G + np.dot(a_KAU.T, np.dot(L_eff, alpha_new - alpha_gov_new) )
        
        d = 1-tax_rates['dividends']
        

        
        alpha_eff = alpha_new*d + alpha_gov_new*(1-d)
        print(alpha_eff)
        
        x = np.dot(L_eff, alpha_eff)*Y
        
        x = np.reshape(x, (len(x),1)) 
        q = x/KAU_price
        
        
        
        
        params = self.create_par_short(seed, Y, GDP_RoW, RoW_price, c_RoW, alpha_RoW, n_sector, C, delta, gamma, cap_ut,cap_target_speed,  KAU_price, wages, labor_tech_coeff, KAU_from_RoW, n_households, lambdaT, csi, alpha, H_from_RoW, alpha_gov, Gov_from_RoW,  tax_rates, ub_frac, transfer_frac)

        if(c_RoW >0):
            params['RoW_var']['GDP'] = (A_d + B_d*d)*Y/c_RoW
        else:
            params['RoW_var']['GDP'] = 0

        params['localKAUs_var']['tech_coeff'] = []
        params['localKAUs_var']['commodities_stock'] = []
        params['localKAUs_var']['capital_coeff'] = []
        params['localKAUs_var']['capital_stocks'] = []
        params['localKAUs_var']['capital_depreciation'] = []
        params['localKAUs_var']['target_capacity_utilization'] = []
        params['localKAUs_var']['my_seller'] = []

        for j in range(n_sector):
            
            params['localKAUs_var']['previous_demand'][j] = (1-C[j,j])*q[j,0]
            params['localKAUs_var']['markup'][j] = 1/(np.sum(C[:,j])) - 1
            
            theta = tax_rates['VAT'][j]
            earnings = (1-sum(A_mod[:,j]) - theta/(1+theta) )*x[j] - wages[j]*labor_tech_coeff[j]*q[j]
            
            params['localKAUs_var']['net_earnings'][j] = (1-tax_rates['corporate'])*earnings
            
            tech_coeff_dict = {}
            stock_dict = {}
            capital_tech_coeff_dict = {}
            capital_stock_dict = {}
            depreciation_dict = {}
            cap_util_dict = {}
            
            my_seller_dict = {}
            for i,commodity in enumerate(params['commodities_list']):
                tech_coeff_dict[commodity] = C[i,j]
                stock_dict[commodity] = C[i,j]*q[j,0]*n_indep_periods
                capital_tech_coeff_dict[commodity] = gamma[i,j]
                capital_stock_dict[commodity] = 0
                if(cap_ut[i,j]>0):
                    capital_stock_dict[commodity] = gamma[i,j]*q[j,0]/cap_ut[i,j]
                depreciation_dict[commodity] = delta[i,j]
                cap_util_dict[commodity] = cap_ut[i,j]
                my_seller_dict[commodity] = i
            
            params['localKAUs_var']['tech_coeff'].append(tech_coeff_dict)
            params['localKAUs_var']['commodities_stock'].append(stock_dict)
            params['localKAUs_var']['capital_coeff'].append(capital_tech_coeff_dict)
            params['localKAUs_var']['capital_stocks'].append(capital_stock_dict)
            params['localKAUs_var']['capital_depreciation'].append(depreciation_dict)
            params['localKAUs_var']['target_capacity_utilization'].append(cap_util_dict)
            params['localKAUs_var']['my_seller'].append(my_seller_dict)
        
    
        KAU_workers = labor_tech_coeff*q
        for k in range(n_sector):
            
            if(KAU_workers[k] - int(KAU_workers[k])>0):
                KAU_workers[k] = int(KAU_workers[k])+1
            else:
                KAU_workers[k] = int(KAU_workers[k])
        
        KAU_effective_wage = wages*labor_tech_coeff*q/KAU_workers
        
        params['Households_var']['employer'] = [-1 for h in range(n_households)]
        
        
        labor_income = np.sum(wages*labor_tech_coeff*q)
        average_wage = 0
        if(np.sum(KAU_workers)>0):
            average_wage = labor_income/np.sum(KAU_workers)
        
        print('w average', average_wage)
        transfer_h = transfer_frac*average_wage
        
        ub_h = ub_frac*average_wage
        
        total_profits = np.sum(params['localKAUs_var']['net_earnings'])

        income_common_part = (1- tax_rates['dividends'])*total_profits/n_households + transfer_h
        
        
        params['Households_var']['wealth'] = np.array([lT*income_common_part for lT in lambdaT])

        N_m = 0
        N_M = 0   
                
        for k in range(n_sector):
            

            N_M += int(KAU_workers[k,0])
            print(N_M)
            for h in range(N_m, N_M):
                
                params['Households_var']['employer'][h] = k                
                params['Households_var']['wealth'][h] += lambdaT[h]*KAU_effective_wage[k]*(1-tax_rates['labor'][math.inf]) 
                
            N_m = N_M 
            
        for h in range(N_M, n_households):
            
            params['Households_var']['wealth'][h] += lambdaT[h]*(ub_h) 

            
        params['localKAUs_var']['prices'] = []
        for i in range(n_sector):
            
            params['localKAUs_var']['prices'].append({comm: KAU_price[j,0]*(1-KAU_from_RoW[j,i]) + RoW_price[j]*KAU_from_RoW[j,i] for j,comm in enumerate(params['commodities_list'])})
 
            
        labor_tax = labor_income*tax_rates['labor'][math.inf]
        dvd_tax = total_profits*tax_rates['dividends']
        corporate_tax = total_profits*tax_rates['corporate']/(1-tax_rates['corporate'])
        VAT = np.sum(tax_rates['VAT']/(1+tax_rates['VAT']) * x)
        
        params['Gov_var']['liquidity'] = labor_tax + dvd_tax + corporate_tax + VAT
        
        print('test',params['Gov_var']['liquidity'] - transfer_h*n_households - ub_h*(n_households-N_M), (1-d)*Y)
        
        
        
        
        return params
    
    
    def create_par_short(self,seed, Y, GDP_RoW, RoW_price, c_RoW, alpha_RoW, n_sector, C, delta, gamma, cap_ut , cap_target_speed, KAU_price, wages, labor_tech_coeff , KAU_from_RoW, n_households, lambdaT, csi, alpha, H_from_RoW, alpha_gov ,Gov_from_RoW, tax_rates, ub_frac, transfer_frac, rat_threshold = 1, n_indep_periods =100, shocks = {'active': False}):

        params = {}
        
        params['shocks'] = shocks
        
        params['seed'] = seed      
        params['KAUtype'] = 'price'        
        params['round_number'] = 5
                
        
        params['nFirms'] = n_sector
        params['nKAUs'] = n_sector
        params['nHouseholds'] = n_households

        params['activities_list'] = ['A' + str(i) for i  in range(1,n_sector+1)]      
        params['commodities_list'] = ['P' + str(i) for i  in range(1,n_sector+1) ]
        
        params['KAUactivity_list'] = ['A' + str(i) for i  in range(1,n_sector+1)]   
        params['KAUcommodity_list'] = ['P' + str(i) for i  in range(1,n_sector+1) ]
        

        params['KAUFirm_list'] =  list(np.arange(n_sector))
        
        params['RoW_var'] = {}
        params['RoW_var']['consumption_shares'] = { comm: alpha_RoW[i] for (i,comm) in enumerate(params['commodities_list'])}
        params['RoW_var']['liquidity'] = 100*GDP_RoW
        #params['RoW_var']['GDP'] = GDP_RoW
        params['RoW_var']['consumption2GDP'] = c_RoW
        params['RoW_var']['suppliers_weights'] = {comm: [1] for comm in params['commodities_list']}
        params['RoW_var']['prices'] = {comm: RoW_price[i] for i,comm in enumerate(params['commodities_list'])}
        
        params['Gov_var'] = {}
        #params['Gov_var']['liquidity']
        params['Gov_var']['tax_rates'] = tax_rates.copy()
        params['Gov_var']['tax_rates']['VAT'] = {comm: tax_rates['VAT'][i] for (i,comm) in enumerate(params['commodities_list'])}
        params['Gov_var']['consumption_shares'] = { comm: alpha_gov[i] for (i,comm) in enumerate(params['commodities_list'])}
        params['Gov_var']['consumption_from_RoW'] = {comm: Gov_from_RoW[i] for i,comm in enumerate(params['commodities_list'])} 
        params['Gov_var']['ub_fraction'] = ub_frac
        params['Gov_var']['transfer_fraction'] = transfer_frac

        params['nBanks'] = 1
        
        params['Banks_var'] = {}        
        params['Banks_var']['reserves'] = [1000]
        params['Banks_var']['CAR'] = [0]
        params['Banks_var']['interest_rate'] = [0.0]        
        params['Banks_var']['threshold'] = [-1]
        params['Banks_var']['repayment_time'] = [1]    



        params['Firms_var'] = {}
        
        params['Firms_var']['loans'] = [list() for i in range(n_sector)]
        params['Firms_var']['wealth'] = [1000*Y for i in range(n_sector)]
        
        # Household initialization
        
        params['Households_var'] = {}
        
        params['Households_var']['property_shares'] = [[1/n_households for f in range(n_sector)] for h in range(n_households)]        
        params['Households_var']['dividends'] =  [0 for lT in lambdaT]
        params['Households_var']['wealth2income_target'] =  [lT for lT in lambdaT]
        params['Households_var']['csi'] =  [csi_h for csi_h in csi]
        
        params['Households_var']['consumption_shares'] = []
        params['Households_var']['my_seller'] = []
        
        for h in range(n_households):
            
            consshare_dict = {}
            my_seller_dict = {}
            for i,commodity in enumerate(params['commodities_list']):
                consshare_dict[commodity] = alpha[i]
                my_seller_dict[commodity] = i
            
            params['Households_var']['consumption_shares'].append(consshare_dict)
            params['Households_var']['my_seller'].append(my_seller_dict)
        
            
        params['Households_var']['rationing_threshold'] = [rat_threshold for h in range(n_households)]
        params['Households_var']['field_of_view']= [1 for h in range(n_households)]
        params['Households_var']['memory_loss'] = [0 for h in range(n_households)]
        params['Households_var']['opportunism_degree'] = [0 for h in range(n_households)]        
        params['Households_var']['consumption_from_RoW'] = [{comm: H_from_RoW[i] for i,comm in enumerate(params['commodities_list'])} for h in range(n_households)]
        
        params['localKAUs_var'] = {}
        

                
        params['localKAUs_var']['previous_demand'] = np.zeros(n_sector)
        params['localKAUs_var']['markup'] = np.zeros(n_sector)
        params['localKAUs_var']['net_earnings'] = np.zeros(n_sector)
        
        params['localKAUs_var']['rationing_threshold'] = [rat_threshold for k in range(n_sector)]
        params['localKAUs_var']['field_of_view']= [1 for k in range(n_sector)]
        params['localKAUs_var']['memory_loss'] = [0 for k in range(n_sector)]
        params['localKAUs_var']['opportunism_degree'] = [0 for k in range(n_sector)]
        params['localKAUs_var']['wage_offer'] = np.reshape(wages, (n_sector,))
        params['localKAUs_var']['labor_tech_coeff'] = np.reshape(labor_tech_coeff, (n_sector,))
        params['localKAUs_var']['consumption_from_RoW'] = [{comm: KAU_from_RoW[i,j] for i,comm in enumerate(params['commodities_list'])} for j in range(n_sector)]
        
        # prod rule and intermediate goods demand parameters
        params['localKAUs_var']['independence_periods'] = np.array([0 for i in range(n_sector)])
        params['localKAUs_var']['target_speed'] = np.array([0 for i in range(n_sector)])
        params['localKAUs_var']['capital_target_speed'] = np.array([v for v in cap_target_speed])
        
        # price dynamics
        params['localKAUs_var']['markup_speed'] = np.array([0 for i in range(n_sector)])
        
        return params
    
    
    def create_initcond_Baseline(self, seed, Y, n_firms, n_KAUs4sector, C, wages, labor_tech_coeff , n_households, lambdaT, csi, alpha,  households_sellers, KAUs_sellers, rat_threshold, n_indep_periods =100, shocks = {'active': False}):

        params = {}
        
        params['shocks'] = shocks
        
        params['seed'] = seed      
        params['KAUtype'] = 'price'        
        params['round_number'] = 5
                
        n_sectors = len(n_KAUs4sector)        
        n_KAUs = sum(n_KAUs4sector)
        
        params['nFirms'] = n_firms  
        params['nKAUs'] = n_KAUs
        params['nHouseholds'] = n_households

        params['activities_list'] = ['A' + str(i) for i  in range(1,n_sectors+1)]      
        params['commodities_list'] = ['P' + str(i) for i  in range(1,n_sectors+1) ]
        
        
        params['KAUactivity_list'] = []
        params['KAUcommodity_list'] = []
        
        for i,nk in enumerate(n_KAUs4sector):

            params['KAUactivity_list'] += ['A' + str(i+1) for k  in range(nk)]
            params['KAUcommodity_list'] += ['P' + str(i+1) for k in range(nk)]
            
       
        params['KAUFirm_list'] =  list(np.arange(n_KAUs))


        params['nBanks'] = 1
        
        params['Banks_var'] = {}        
        params['Banks_var']['reserves'] = [1000]
        params['Banks_var']['CAR'] = [0]
        params['Banks_var']['interest_rate'] = [0.0]
        
        params['Banks_var']['threshold'] = [-1]
        params['Banks_var']['repayment_time'] = [1]    



        params['Firms_var'] = {}
        
        params['Firms_var']['loans'] = [list() for i in range(n_firms)]  

        
        Total_firms_wealth = n_KAUs*1000*(np.sum(lambdaT) + 1) # every sector has 1000 the households wealth to avoid rationing
        
                
        # in order to have KAUs with equal wealth
        
        params['Firms_var']['wealth'] = [Total_firms_wealth/n_KAUs*params['KAUFirm_list'].count(i) for i in range(n_firms)]
        
        # Household initialization
        
        params['Households_var'] = {}
        
        params['Households_var']['property_shares'] = [[1/n_households for f in range(n_firms)] for h in range(n_households)]

        
        params['Households_var']['dividends'] =  [0 for lT in lambdaT]
        params['Households_var']['wealth2income_target'] =  [lT for lT in lambdaT]
        params['Households_var']['csi'] =  [csi_h for csi_h in csi]
        
        params['Households_var']['consumption_shares'] = []
        params['Households_var']['my_seller'] = []
        
        for h in range(n_households):
            
            consshare_dict = {}
            my_seller_dict = {}
            for i,commodity in enumerate(params['commodities_list']):
                consshare_dict[commodity] = alpha[i,h]
                my_seller_dict[commodity] = households_sellers[i,h]
            
            params['Households_var']['consumption_shares'].append(consshare_dict)
            params['Households_var']['my_seller'].append(my_seller_dict)
        
            
        params['Households_var']['rationing_threshold'] = [rat_threshold for h in range(n_households)]
        params['Households_var']['field_of_view']= [1 for h in range(n_households)]
        params['Households_var']['memory_loss'] = [0 for h in range(n_households)]
        params['Households_var']['opportunism_degree'] = [0 for h in range(n_households)]
            
        # local KAUs initialization
        
        params['localKAUs_var'] = {}
        
        C_new = np.zeros((n_KAUs,n_KAUs))
        
        n_KAUs4sector_cum = np.cumsum(n_KAUs4sector)
        
        for j in range(n_KAUs):
            
            s = 0
            for i in range(n_KAUs):
                
                if(i== n_KAUs4sector_cum[s]):
                    s=s+1
                if(i==KAUs_sellers[s,j]):
                    C_new[i,j] = C[s,j]

        
        alpha_new = np.zeros((n_KAUs,n_households))
        
        for h in range(n_households):
          
            s = 0
            for i in range(n_KAUs):
                
                if(i== n_KAUs4sector_cum[s]):
                    s=s+1
                if(i == households_sellers[s,h]):
                        
                    alpha_new[i,h] = alpha[s,h]
                        
                        
        alpha_eff = np.reshape(alpha_new.mean(axis=1), (n_KAUs, 1))
        
        L_new = np.linalg.inv(np.eye(n_KAUs)-C_new)    
        q2Y = L_new.dot(alpha_eff)
        q = q2Y*Y
      
        params['localKAUs_var']['tech_coeff'] = []
        params['localKAUs_var']['commodities_stock'] = []
        params['localKAUs_var']['my_seller'] = []
        
        
        params['localKAUs_var']['previous_demand'] = np.zeros(n_KAUs)
        params['localKAUs_var']['markup'] = np.zeros(n_KAUs)
        params['localKAUs_var']['earnings'] = np.zeros(n_KAUs)
        
        for j in range(n_KAUs):
            
            params['localKAUs_var']['previous_demand'][j] = (1-C_new[j,j])*q[j]
            params['localKAUs_var']['markup'][j] = 1/(np.sum(C_new[:,j])) - 1
            params['localKAUs_var']['earnings'][j] = (1-sum(C_new[:,j]) - wages[j]*labor_tech_coeff[j])*q[j]
            
            tech_coeff_dict = {}
            stock_dict = {}
            my_seller_dict = {}
            for i,commodity in enumerate(params['commodities_list']):
                tech_coeff_dict[commodity] = C[i,j]
                stock_dict[commodity] = C[i,j]*q[j,0]*n_indep_periods
                my_seller_dict[commodity] = KAUs_sellers[i,j]
            
            params['localKAUs_var']['tech_coeff'].append(tech_coeff_dict)
            params['localKAUs_var']['commodities_stock'].append(stock_dict)    
            params['localKAUs_var']['my_seller'].append(my_seller_dict)
        
        params['localKAUs_var']['rationing_threshold'] = [rat_threshold for k in range(n_KAUs)]
        params['localKAUs_var']['field_of_view']= [1 for k in range(n_KAUs)]
        params['localKAUs_var']['memory_loss'] = [0 for k in range(n_KAUs)]
        params['localKAUs_var']['opportunism_degree'] = [0 for k in range(n_KAUs)]
        params['localKAUs_var']['wage_offer'] = np.reshape(wages, (n_KAUs,))
        params['localKAUs_var']['labor_tech_coeff'] = np.reshape(labor_tech_coeff, (n_KAUs,))
    
        wage2income = np.sum(wages*labor_tech_coeff*q2Y)
        profit2income = 1- wage2income
    
        KAU_workers = labor_tech_coeff*q
        for k in range(n_KAUs):
            
            if(KAU_workers[k] - int(KAU_workers[k])>0):
                KAU_workers[k] = int(KAU_workers[k])+1
            else:
                KAU_workers[k] = int(KAU_workers[k])
        
        KAU_effective_wage = wages*labor_tech_coeff*q/KAU_workers
        
        params['Households_var']['employer'] = [-1 for h in range(n_households)]
        params['Households_var']['wealth'] = np.array([(lT)/n_households*profit2income*Y for lT in lambdaT])
        
        N_m = 0
        N_M = 0   
        
        
        for k in range(n_KAUs):
            

            N_M += int(KAU_workers[k,0])
            print(N_M)
            for h in range(N_m, N_M):
                
                params['Households_var']['employer'][h] = k                
                params['Households_var']['wealth'][h] += lambdaT[h]*KAU_effective_wage[k]
                
            N_m = N_M 

            
        params['localKAUs_var']['prices'] = {} #{'P1':1., 'P2':1., 'P3':1.}
        for commodity in params['commodities_list']:
            params['localKAUs_var']['prices'][commodity] = 1.  
            
        
        
        # prod rule and intermediate goods demand parameters
        params['localKAUs_var']['independence_periods'] = np.array([0 for i in range(n_KAUs)])
        params['localKAUs_var']['target_speed'] = np.array([0 for i in range(n_KAUs)])
        
        # price dynamics
        params['localKAUs_var']['markup_speed'] = np.array([0 for i in range(n_KAUs)])

        
        return params,q, KAU_workers
      



    def run_single_realization(self, params, n_steps):
        
        m = md.MyModel(params)        
        results = m.run(n_steps)
        
        self.model = m
        
        return results
 
    
    def run_single_scenario_multiple_seeds(self, params,  seeds, n_steps):
        
        # results = {}
        
        # for s in seeds:
            
        #     params['seed'] = s
        #     results[s] = self.run_single_realization(params, n_steps)
        results = self.run_single_realization(params, n_steps)
        
        return results
    
    def execute_single_scenario(self, params, n_steps, folder_path):

        
        results = self.run_single_realization(params, n_steps)
       
        self.figures_single_sim(results, folder_path)
        self.write_report_file(results, ['',''], [1], [''], folder_path)
        
        file = open(folder_path + '/Data', 'wb')
        pk.dump(results, file)
        file.close()
        
        return results


    
    def execute_single_scenario_with_sol(self, params, n_steps, folder_path):

        
        self.find_fixedpoint_baseline(params)
        
        results = self.run_single_realization(params, n_steps)
       
        self.figures_single_sim_fixedpoint(results, folder_path)
        self.write_report_file(results, ['',''], [1], [''], folder_path)
        
        file = open(folder_path + '/Data', 'wb')
        pk.dump(results, file)
        file.close()
        
        return results    
    
    

    def copy_list(self, list_original):
        
        copy_list = []
        
        for el in list_original:
            if(type(el) == dict):
                copy_list.append(self.copy_dictionary(el))
            elif(type(el) == list):
                copy_list.append(self.copy_list(el))
            else:
                try:
                    copy_list.append(el.copy())
                except:
                    copy_list.append(el)
                    
        return copy_list 

    def copy_dictionary(self, dictionary):
        
        copy_dict = {}
        for k in dictionary.keys():
            
            if(type(dictionary[k]) == dict):
                copy_dict[k] = self.copy_dictionary(dictionary[k])
            elif(type(dictionary[k]) == list):
                copy_dict[k] = self.copy_list(dictionary[k])
            else:
                try:
                    copy_dict[k] = dictionary[k].copy()
                except:
                    copy_dict[k] = dictionary[k]
                
        return copy_dict


    def construct_par2iter(self, params, var2iter_labels, values2iter, operation):
        
        
        par2iter = []
        
        par0 = np.array([])
        
        if(len(var2iter_labels)==1):
                 par0 = params[var2iter_labels[0]]
        elif(len(var2iter_labels)==2):
                 par0 = params[var2iter_labels[0]][var2iter_labels[1]]
        else:
            print('too many labels in var2iter')
            return 
        
        for i in range(len(values2iter)):
            if(operation == '*'):
                par2iter.append(values2iter[i]*par0)
            else:
                par2iter.append(values2iter[i] + par0)
                
        return par2iter
    
    

   
    def run_multiple_scenarios(self, params, var2iter_labels, values2iter, operation, seeds, n_steps, scenarios_list):
               
        results = {}
        
        par2iter = self.construct_par2iter(params, var2iter_labels, values2iter, operation)
        
        for i in range(len(par2iter)):
            
            par = self.copy_dictionary(params)
            
            if(len(var2iter_labels)==1):
                par[var2iter_labels[0]] = par2iter[i]
            elif(len(var2iter_labels)==2):
                par[var2iter_labels[0]][var2iter_labels[1]] = par2iter[i]
            else:
                print('too many labels in var2iter')
                return
            
            results[scenarios_list[i]] = self.run_single_scenario_multiple_seeds(par, seeds, n_steps)
                
            
 
        return results, par2iter
    
    def run_multiple_scenariosV2(self, params, n_steps, scenarios_list, folder_path,MACRO = 0):
        
        results = {}
        for i,p in enumerate(params):
            results[scenarios_list[i]] = self.run_single_realization(p, n_steps)

        if(MACRO):
            self.figures_more_sim_MACRO(results, folder_path, scenarios_list)
        else:
            self.figures_more_sim(results, folder_path, scenarios_list)
        
        
        #self.write_report_file(results, var2iter_labels, par2iter, scenarios_list, folder_path)
        
        
        file = open(folder_path + '/Data', 'wb')
        pk.dump(results, file)
        file.close()
        
        return results

        
    def execute_multiple_scenarios(self, params, KAU_type, seed, n_steps, scenarios_list, var2iter_labels, values2iter, operation, folder_path):
        
        
        results, par2iter = self.run_multiple_scenarios(params, var2iter_labels, values2iter, operation, seed, n_steps, scenarios_list)
        
        
        self.figures_more_sim(results, folder_path, scenarios_list)
        self.write_report_file(results, var2iter_labels, par2iter, scenarios_list, folder_path)
        
        
        file = open(folder_path + '/Data', 'wb')
        pk.dump(results, file)
        file.close()

        
        
        
        return results, par2iter
    
    
    def figures_more_sim_MACRO(self, results, folder_path, scenarios_list):


        if(type(results) != dict):
            return 'Error: results variable has to be a dict'

                
        if(not os.path.isdir(folder_path)):
            os.makedirs(folder_path)
       
        
        for key in results[scenarios_list[0]].variables.keys():
            
            if(not os.path.isdir(folder_path + '/' + key)):
                os.mkdir(folder_path + '/' + key)
                os.mkdir(folder_path + '/' + key + '/pyfig')
            
            if(key == 'MyModel'):
                for cols in results[scenarios_list[0]].variables[key].columns:
                    f= plt.figure()
                    for sc in scenarios_list:
                        
                        x = np.array(results[sc].variables[key][cols])
                        if('wealth' in cols or 'stock' in cols ):
                            plt.plot(x)
                        else:
                            plt.plot(x[1:])
                        
                    plt.ylabel(cols)
                    plt.xlabel('Time')
                    plt.legend(scenarios_list)
                    plt.grid()
                    f.savefig(folder_path + '/'+ key +'/' + cols + '.png')
                    file4fig = open(folder_path + '/'+ key +'/pyfig/' + cols, 'wb')
                    pk.dump(f, file4fig)
                    file4fig.close()
                    plt.close()    


    def figures_more_sim(self, results, folder_path, scenarios_list):
        
        if(type(results) != dict):
            return 'Error: results variable has to be a dict'

                
        if(not os.path.isdir(folder_path)):
            os.makedirs(folder_path)
       
        
        for key in results[scenarios_list[0]].variables.keys():
            
            if(not os.path.isdir(folder_path + '/' + key)):
                os.mkdir(folder_path + '/' + key)
                os.mkdir(folder_path + '/' + key + '/pyfig')
            
            if(key == 'MyModel'):
                for cols in results[scenarios_list[0]].variables[key].columns:
                    f= plt.figure()
                    for sc in scenarios_list:
                        
                        x = np.array(results[sc].variables[key][cols])
                        if('wealth' in cols or 'stock' in cols ):
                            plt.plot(x)
                        else:
                            plt.plot(x[1:])
                        
                    plt.ylabel(cols)
                    plt.xlabel('Time')
                    plt.legend(scenarios_list)
                    plt.grid()
                    f.savefig(folder_path + '/'+ key +'/' + cols + '.png')
                    file4fig = open(folder_path + '/'+ key +'/pyfig/' + cols, 'wb')
                    pk.dump(f, file4fig)
                    file4fig.close()
                    plt.close()
            elif(key != 'Household'):
                print(key)
                ids = np.unique(results[scenarios_list[0]].variables[key].index.get_level_values(0))
                for i in ids:
                    for cols in results[scenarios_list[0]].variables[key].columns:
                        f= plt.figure()
                        for sc in scenarios_list:
                            key2 = key
                            if(not key in results[sc].variables.keys()):
                                key2 = list(results[sc].variables.keys() - results[0].variables.keys())[0]
                            x = np.array(results[sc].variables[key2][cols].loc[i,:])
                            if('wealth' in cols or 'stock' in cols ):
                                plt.plot(x)
                            else:
                                plt.plot(x[1:])
                        plt.ylabel(cols)
                        plt.xlabel('Time')
                        plt.legend(scenarios_list)
                        plt.grid()
                        f.savefig(folder_path + '/'+ key +'/' + str(i) + cols + '.png') 
                        file4fig = open(folder_path + '/'+ key +'/pyfig/' + str(i) + cols, 'wb')
                        pk.dump(f, file4fig)
                        file4fig.close()
                        plt.close()



    def write_report_file(self, results, var2iter_labels, par2iter, scenarios_list, folder_path):
        
        
        report_file = open(folder_path + '/ReadMe.txt', 'w')
        
        report_file.write('Simulation where ' + var2iter_labels[0] + 'are changed\n')
        report_file.write('Values of the parameters are:\n')
        for i in range(len(par2iter)):
            report_file.write(var2iter_labels[0] + ' ' + var2iter_labels[1] + ' ' + str(i) + ':\n')
            try:
                for j in range(len(par2iter[i])):
                    report_file.write('\t' + str(par2iter[i][j]) + '\n')
            except:
                report_file.write('\t' + str(par2iter[i]) + '\n')
                                  
        report_file.write('Scenarios names are:\n')
        
        
        for i,sc in enumerate(scenarios_list):
            report_file.write('\t' + str(i) + ': ' + sc + '\n')
            
        for i in range(len(par2iter)):
            
            try:
                report_file.write('Scenario n. ' + str(i) + 'started in: ' + results[scenarios_list[i]]['info']['time_stamp'] + '\n')
                report_file.write('Execution time: ' + results[scenarios_list[i]]['info']['run_time'] + '\n\n')
            except:
                report_file.write('Scenario n. ' + str(i) + 'started in: ' + results['info']['time_stamp'] + '\n')
                report_file.write('Execution time: ' + results['info']['run_time'] + '\n\n')    
            
            
        report_file.close()
        
        return
        
    

      
    
    def figures_single_sim(self, results, folder_path):
        
        
        if(not os.path.isdir(folder_path)):
            os.makedirs(folder_path)
       
        
        for key in results.variables.keys():
            
            if(not os.path.isdir(folder_path + '/' + key)):
                os.mkdir(folder_path + '/' + key)
                
            if(not(os.path.isdir(folder_path + '/'+ key +'/pyfig'))):
                os.mkdir(folder_path + '/'+ key +'/pyfig')
                
                
            if(key == 'MyModel'):                
                for cols in results.variables[key].columns:
                    x = np.array(results.variables[key][cols])
                    f= plt.figure()
                    if('wealth' in cols or 'stock' in cols ):
                        plt.plot(x)
                    else:
                        plt.plot(x[1:])
                    plt.ylabel(cols)
                    plt.xlabel('Time')
                    plt.grid()
                    f.savefig(folder_path + '/'+ key +'/' + cols + '.png')
                    file4fig = open(folder_path + '/'+ key +'/pyfig/' + cols, 'wb')
                    pk.dump(f, file4fig)
                    file4fig.close()
                    plt.close()
            elif(key != 'Household'):
                
                ids = np.unique(results.variables[key].index.get_level_values(0))
                for i in ids:
                    for cols in results.variables[key].columns:
                        x = np.array(results.variables[key][cols].loc[i,:])
                        f = plt.figure()
                        if('wealth' in cols or 'stock' in cols ):
                            plt.plot(x)
                        else:
                            plt.plot(x[1:])
                        plt.ylabel(cols)
                        plt.xlabel('Time')
                        plt.grid()
                        f.savefig(folder_path + '/'+ key +'/' + str(i) + cols + '.png')
                        file4fig = open(folder_path + '/'+ key +'/pyfig/' + str(i) + cols, 'wb')
                        pk.dump(f, file4fig)
                        file4fig.close()
                        plt.close()
                        plt.close()




    
    def find_fixedpoint_baseline(self, params):
        
        solution = {}
        
        
        # ricavare matrice C, alpha, lambdaT
        
        C = np.zeros((params['nKAUs'],params['nKAUs']))
        
        for j,v in enumerate(params['localKAUs_var']['tech_coeff']):        
            for i,val in enumerate(v.values()):
                C[i,j] = val
        
        alpha = np.array([val for val in params['Households_var']['consumption_shares'][0].values()]).reshape((params['nKAUs'],1))
        
        lambdaT = params['Households_var']['wealth2income_target'][0]
        
        # ricavare richezza H, dividendi e domande
        
        M_H0 = params['Households_var']['wealth'][0]
        
        DIV0 = np.sum(params['localKAUs_var']['earnings'])
        
        
        qD0 = params['localKAUs_var']['previous_demand']
        qD0_sum = np.sum(qD0)
        
        
        # fixed-point income
        
        Idm = np.eye(params['nKAUs'])
        
        C_hat = np.diag(np.diag(C))
        
            # To avoid matrix inverse calculation
        
        Lalpha = np.linalg.solve(Idm-C, alpha)
        
        
        qD2Y = (Idm-C_hat).dot(Lalpha)
        
        
        Y = M_H0 + DIV0 + qD0_sum
        
        Y = Y/(lambdaT + 1 + np.sum(qD2Y))
        
        
        # 
        
        
        M_F0 = np.sum(params['Firms_var']['wealth'])
     
        
        solution['MyModel'] = {}
        
        solution['MyModel']['GDP'] = Y
        solution['MyModel']['Consumption'] = Y
        
        for i,com in enumerate(params['commodities_list']):
            
            solution['MyModel']['Real production of '+ com] = Y*Lalpha[i]
        
        solution['MyModel']['Nominal Production'] = Y*np.sum(Lalpha)

        
        solution['MyModel']['Households liquidity'] = lambdaT*Y
        
        solution['MyModel']['Firms liquidity'] = M_F0 - lambdaT*Y + M_H0
    
        
    
        solution['LocalKAU_price'] = {}
    
        I = np.zeros((params['nKAUs'],params['nKAUs']))
        
        qD = qD2Y*Y
        
        for j in range(params['nKAUs']):
            for i,com in enumerate(params['commodities_list']):
                
                I[i,j] = params['localKAUs_var']['commodities_stock'][j][com]
        
        for i in range(params['nKAUs']):
            
            I[i,i] += qD0[i] - qD[i]
        
        for i,com in enumerate(params['commodities_list']):
            solution['LocalKAU_price'][com + ' stock'] = I[i,:]
        
        
        self.solution = solution
        
        #return solution

    def find_fixedpoint_baseline_N_Households(self, params):
        
        solution = {}
        
        
        # ricavare matrice C, alpha, lambdaT
        
        C = np.zeros((params['nKAUs'],params['nKAUs']))
        
        for j,v in enumerate(params['localKAUs_var']['tech_coeff']):        
            for i,val in enumerate(v.values()):
                C[i,j] = val
        
        
        alpha = np.zeros((params['nKAUs'],params['nHouseholds']))
        
        for h in range(params['nHouseholds']):
            
            for i,val in enumerate(params['Households_var']['consumption_shares'][h].values()):
                
                alpha[i,h] = val
        
        
        #da correggere       
        sh_matrix = np.array(params['Households_var']['property_shares'])
        
        sh = sh[:,0]
        sh = sh.reshape((params['nKAUs'],1))

        alpha_eq = alpha.dot(sh)   #reshape((params['nKAUs'],1)
        
        lambdaT = params['Households_var']['wealth2income_target']
        
        lambdaT_eq = np.sum(lambdaT*sh)
        
        # ricavare richezza H, dividendi e domande
        
        M_H0 = np.sum(params['Households_var']['wealth'])
        
        DIV0 = np.sum(params['localKAUs_var']['earnings'])
        
        
        qD0 = params['localKAUs_var']['previous_demand']
        qD0_sum = np.sum(qD0)
        
        
        # fixed-point income
        
        Idm = np.eye(params['nKAUs'])
        
        C_hat = np.diag(np.diag(C))
        
            # To avoid matrix inverse calculation
        
        Lalpha = np.linalg.solve(Idm-C, alpha_eq)
        
        
        qD2Y = (Idm-C_hat).dot(Lalpha)
        
        
        Y = M_H0 + DIV0 + qD0_sum
        
        Y = Y/(lambdaT_eq + 1 + np.sum(qD2Y))
        
        
        # 
        
        
        M_F0 = np.sum(params['Firms_var']['wealth'])
     
        
        solution['MyModel'] = {}
        
        solution['MyModel']['GDP'] = Y
        solution['MyModel']['Consumption'] = Y
        
        for i,com in enumerate(params['commodities_list']):
            
            solution['MyModel']['Real production of '+ com] = Y*Lalpha[i]
        
        solution['MyModel']['Nominal Production'] = Y*np.sum(Lalpha)

        
        solution['MyModel']['Households liquidity'] = lambdaT_eq*Y
        
        solution['MyModel']['Firms liquidity'] = M_F0 - lambdaT_eq*Y + M_H0
    
        
    
        solution['LocalKAU_price'] = {}
    
        I = np.zeros((params['nKAUs'],params['nKAUs']))
        
        qD = qD2Y*Y
        
        for j in range(params['nKAUs']):
            for i,com in enumerate(params['commodities_list']):
                
                I[i,j] = params['localKAUs_var']['commodities_stock'][j][com]
        
        for i in range(params['nKAUs']):
            
            I[i,i] += qD0[i] - qD[i]
        
        for i,com in enumerate(params['commodities_list']):
            solution['LocalKAU_price'][com + ' stock'] = I[i,:]
        
        
        self.solution = solution



    def figures_single_sim_fixedpoint(self, results, folder_path):
        
                
        if(not os.path.isdir(folder_path)):
            os.makedirs(folder_path)
       
        
        for key in results.variables.keys():
            
            if(not os.path.isdir(folder_path + '/' + key)):
                os.mkdir(folder_path + '/' + key)
                
            if(not(os.path.isdir(folder_path + '/'+ key +'/pyfig'))):
                os.mkdir(folder_path + '/'+ key +'/pyfig')
                
                
            if(key == 'MyModel'):                
                for cols in results.variables[key].columns:
                    
                    x = np.array(results.variables[key][cols])
                    f= plt.figure()
                    if('wealth' in cols or 'stock' in cols ):
                        plt.plot(x)
                    else:
                        plt.plot(x[1:])
                        
                    if(cols in self.solution['MyModel'].keys()):
                        
                        sol = self.solution['MyModel'][cols]
                        if('wealth' in cols or 'stock' in cols ):
                            plt.plot(np.ones(len(x))*sol, '--')
                        else:
                            plt.plot(np.ones(len(x)-1)*sol, '--')
                                                
                    plt.ylabel(cols)
                    plt.xlabel('Time')
                    plt.grid()
                    f.savefig(folder_path + '/'+ key +'/' + cols + '.png')
                    file4fig = open(folder_path + '/'+ key +'/pyfig/' + cols, 'wb')
                    pk.dump(f, file4fig)
                    file4fig.close()
                    plt.close()
            else:
                
                ids = np.unique(results.variables[key].index.get_level_values(0))
                for ni,i in enumerate(ids):
                    for cols in results.variables[key].columns:
                        x = np.array(results.variables[key][cols].loc[i,:])
                        f = plt.figure()
                        if('wealth' in cols or 'stock' in cols ):
                            plt.plot(x)
                        else:
                            plt.plot(x[1:])
                            

                        if(cols in self.solution['LocalKAU_price'].keys()):
                            
                            sol = self.solution['LocalKAU_price'][cols][ni]

                            if('wealth' in cols or 'stock' in cols ):
                                plt.plot(np.ones(len(x))*sol, '--')
                            else:
                                plt.plot(np.ones(len(x)-1)*sol, '--')
                                
                                
                        plt.ylabel(cols)
                        plt.xlabel('Time')
                        plt.grid()
                        f.savefig(folder_path + '/'+ key +'/' + str(i) + cols + '.png')
                        file4fig = open(folder_path + '/'+ key +'/pyfig/' + str(i) + cols, 'wb')
                        pk.dump(f, file4fig)
                        file4fig.close()
                        plt.close()
                        plt.close()









    def save_initcond_MASAM_01(self, params, name):
        
        
        wb = opxl.Workbook()
        
        sheet = wb.active
    
        wb.remove(sheet)
        
        General_sheet = wb.create_sheet('General')
        
        General_sheet['A1']= 'Number of Firms'
        General_sheet['B1'] = params['nFirms']
        
        General_sheet['A2'] = 'Number of LocalKAUs'
        General_sheet['B2'] = params['nKAUs']
        
        General_sheet['A3'] = 'Number of Households'
        General_sheet['B3'] = params['nHouseholds']
        
        
        General_sheet['A4'] = 'Activities list'
        for i,a in enumerate(params['activities_list']):
            General_sheet.cell(row = 4, column=i+2, value=a)
        
        General_sheet['A5'] = 'Products list'
        for i,a in enumerate(params['commodities_list']):
            General_sheet.cell(row = 5, column=i+2, value=a)  
        
        
        Households_sheet = wb.create_sheet('Households')
        
        Households_sheet['A2'] = 'wealth'
        Households_sheet['A3'] = 'dividends'
        Households_sheet['A4'] = 'wealth2income_target'
        Households_sheet['A5'] = 'csi'
    
        for i in range(params['nHouseholds']):
            Households_sheet.cell(row = 1, column=i+2, value='H'+str(i+1))
            Households_sheet.cell(row = 2, column=i+2, value= params['Households_var']['wealth'][i])
            Households_sheet.cell(row = 3, column=i+2, value= params['Households_var']['dividends'][i])
            Households_sheet.cell(row = 4, column=i+2, value= params['Households_var']['wealth2income_target'][i])
            Households_sheet.cell(row = 5, column=i+2, value= params['Households_var']['csi'][i])
    
        Households_cons_shares_sheet = wb.create_sheet('Households_cons_shares')
        
        for i,comm in enumerate(params['commodities_list']):
            Households_cons_shares_sheet.cell(row=i+2, column=1, value= comm)
            
        for h in range(params['nHouseholds']):
            
            Households_cons_shares_sheet.cell(row=1, column=h+2, value= 'H'+str(h+1))
            
            for i,comm in enumerate(params['commodities_list']):
                Households_cons_shares_sheet.cell(row=i+2, column=h+2, value= params['Households_var']['consumption_shares'][h][comm])
    
        
    
        
        LocalKAUs_sheet = wb.create_sheet('LocalKAUs')
    
        LocalKAUs_sheet['A2'] = 'activity'
        LocalKAUs_sheet['A3'] = 'products'
        LocalKAUs_sheet['A4'] = 'owner'
        LocalKAUs_sheet['A5'] = 'previous_demand'
        LocalKAUs_sheet['A6'] = 'markup'
        LocalKAUs_sheet['A7'] = 'independence_periods'
        LocalKAUs_sheet['A8'] = 'target_speed'
        LocalKAUs_sheet['A9'] = 'markup_speed'
        
        for k in range(params['nKAUs']):
            LocalKAUs_sheet.cell(row = 1, column=k+2, value='K'+str(k+1))
            LocalKAUs_sheet.cell(row = 2, column=k+2, value= params['KAUactivity_list'][k])
            LocalKAUs_sheet.cell(row = 3, column=k+2, value= params['KAUcommodity_list'][k])
            LocalKAUs_sheet.cell(row = 4, column=k+2, value= params['KAUFirm_list'][k])
            #print(params['localKAUs_var']['previous_demand'][k])
            LocalKAUs_sheet.cell(row = 5, column=k+2, value= params['localKAUs_var']['previous_demand'][k])
            LocalKAUs_sheet.cell(row = 6, column=k+2, value= params['localKAUs_var']['markup'][k])
            LocalKAUs_sheet.cell(row = 7, column=k+2, value= params['localKAUs_var']['independence_periods'][k])
            LocalKAUs_sheet.cell(row = 8, column=k+2, value= params['localKAUs_var']['target_speed'][k])
            LocalKAUs_sheet.cell(row = 9, column=k+2, value= params['localKAUs_var']['markup_speed'][k])
    
    
        LocalKAUs_stocks_sheet = wb.create_sheet('LocalKAUs_stocks')
        
        for k in range(params['nKAUs']):
            LocalKAUs_stocks_sheet.cell(row=1, column = k+2, value= 'K'+str(k+1))
            
            for i,comm in enumerate(params['commodities_list']):
                
                LocalKAUs_stocks_sheet.cell(row=i+2, column = 1, value= comm )
                LocalKAUs_stocks_sheet.cell(row=i+2, column = k+2, value= params['localKAUs_var']['commodities_stock'][k][comm] )
    
    
        LocalKAUs_techcoeff_sheet = wb.create_sheet('LocalKAUs_tech_coeff')
        
        for k in range(params['nKAUs']):
            LocalKAUs_techcoeff_sheet.cell(row=1, column = k+2, value= 'K'+str(k+1))
            
            for i,comm in enumerate(params['commodities_list']):
                
                LocalKAUs_techcoeff_sheet.cell(row=i+2, column = 1, value= comm )
                LocalKAUs_techcoeff_sheet.cell(row=i+2, column = k+2, value= params['localKAUs_var']['tech_coeff'][k][comm] )
    
    
    
        Firms_sheet = wb.create_sheet('Firms')
        
        Firms_sheet['A2'] = 'wealth'
        
        for f in range(params['nFirms']):
            Firms_sheet.cell(row = 1, column = f+ 2, value = 'F'+str(f+1))
            Firms_sheet.cell(row=2, column= f+2, value = params['Firms_var']['wealth'][f])   
    
        wb.save(name)
